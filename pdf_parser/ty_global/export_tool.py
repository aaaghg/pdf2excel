from openpyxl import Workbook
from enum import IntEnum
from pdf_parser.ty_global.order_grouping import GroupType


class TyGlobalColumnSequence(IntEnum):
    Client = 0,
    Supplier = 1,
    OrderNumber = 2,
    ProductNumber = 3,
    ProductName = 4,
    ItemUpcCodeInColorBox = 5,
    UpcCodeInCarton = 6,
    OrderQty = 7,
    PackingWay = 8,
    TotalCartonsPerOrder = 9,
    ProductSpec = 10,
    PackingSpec = 11,


class TyGlobalExportTool:
    def __init__(self, path, supplier, order_number, order_table):
        self.output_file = path
        self.order_table = order_table.data
        self.client = 'Ty Global'
        self.supplier = supplier
        self.order_number = order_number
        self.output_data = self.assemble_data()

    @staticmethod
    def get_only_order_number(order_number):
        return order_number.split('\n')[0]

    def assemble_data(self):
        ret = []
        for packing_group in self.order_table:
            for part_no in self.order_table[packing_group]:
                row = [None]*len(TyGlobalColumnSequence)
                row[TyGlobalColumnSequence.Client.value] = self.client
                row[TyGlobalColumnSequence.Supplier.value] = self.supplier
                row[TyGlobalColumnSequence.OrderNumber.value] = self.order_number
                current_group = self.order_table[packing_group][part_no]
                if current_group.group_type == GroupType.ProductOrder:
                    row[TyGlobalColumnSequence.ProductNumber.value] = part_no
                    row[TyGlobalColumnSequence.ProductName.value] = current_group.product_name
                    row[TyGlobalColumnSequence.ItemUpcCodeInColorBox.value] = current_group.group[current_group.get_item_upc_line_no()][3].split(':')[1].strip()
                    row[TyGlobalColumnSequence.UpcCodeInCarton.value] = current_group.group[current_group.get_carton_upc_line_no()][3].split(':')[1].strip()
                    row[TyGlobalColumnSequence.OrderQty.value] = current_group.group[0][0]
                    ret.append(row[:])
                elif current_group.group_type == GroupType.CasePack:
                    for row in ret:
                        if row[3] in self.order_table[packing_group]:
                            row[TyGlobalColumnSequence.PackingWay.value] = current_group.group[1][3].split(':')[1].strip()
                        else:
                            continue
                elif current_group.group_type == GroupType.ProductSpec:
                    for row in ret:
                        if row[3] in self.order_table[packing_group]:
                            row[TyGlobalColumnSequence.ProductSpec.value] = current_group.get_text()
                        else:
                            continue
                elif current_group.group_type == GroupType.PackingSpec:
                    for row in ret:
                        if row[3] in self.order_table[packing_group]:
                            row[TyGlobalColumnSequence.PackingSpec.value] = current_group.get_text()
                        else:
                            continue
        return ret

    def export(self):
        wb = Workbook()
        title = 'Client	Supplier	Order Number	PRODUCT NUMBER	Product Name	"Item UPC Code\nin Color box"	"UPC Code\n在外箱上"	 Order QTY (units)	Packing way  (Units/carton) 	Total Cartons/per order	Product Size and picture	Packing spec :	驗貨重點'.split('\t')
        start_y = 65
        start_x = 1
        ws = wb.active
        title_y = start_y
        for t in title:
            ws[f'{chr(title_y)}{start_x}'] = t
            title_y = title_y + 1
        print(title)
        start_x = start_x + 1
        order_x = start_x
        for x in range(0, len(self.output_data)):
            for y in range(0, len(self.output_data[x])):
                order_x = start_x + x
                order_y = start_y + y
                ws[f'{chr(order_y)}{order_x}'] = self.output_data[x][y]
        ws[f'H{order_x+1}'] = 'Total Cartons/per order'
        only_order_number = self.get_only_order_number(self.order_number)
        print(only_order_number)
        ws.title = f'{only_order_number} Overall'
        wb.save(self.output_file)


